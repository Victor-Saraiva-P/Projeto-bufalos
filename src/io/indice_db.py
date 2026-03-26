from __future__ import annotations

from collections.abc import Iterable
from decimal import Decimal
from pathlib import Path
from typing import Any

from sqlalchemy import Column, Integer, MetaData, String, Table, create_engine, func, select, update
from sqlalchemy.engine import Engine

from src.config import FAZENDA_COL, INDICE_DB_PATH, LEGACY_INDICE_PATH, NOME_COL, PESO_COL, TAGS_COL
from src.models.indice_linha import IndiceLinha, normalizar_tags


TABLE_NAME = "indice"
metadata = MetaData()
indice_table = Table(
    TABLE_NAME,
    metadata,
    Column("ordem", Integer, nullable=False),
    Column("nome_arquivo", String, primary_key=True, nullable=False),
    Column("fazenda", String, nullable=False),
    Column("peso", String, nullable=False),
    Column("tags", String, nullable=False, default=""),
)


def _engine(db_path: str) -> Engine:
    return create_engine(f"sqlite+pysqlite:///{db_path}")


def _resolver_db_path(db_path: str | None) -> str:
    return db_path or INDICE_DB_PATH


def _resolver_legacy_path(legacy_path: str | None) -> str:
    return legacy_path or LEGACY_INDICE_PATH


def _criar_schema(engine: Engine) -> None:
    metadata.create_all(engine)


def _contar_registros(engine: Engine) -> int:
    with engine.connect() as conn:
        count = conn.execute(select(func.count()).select_from(indice_table)).scalar_one()
    return int(count)


def _carregar_planilha_legada(legacy_path: str) -> list[IndiceLinha]:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        raise ModuleNotFoundError(
            "openpyxl e necessario apenas para migrar a planilha legada para SQLite."
        ) from exc

    workbook = load_workbook(legacy_path, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    workbook.close()

    if not rows:
        return []

    header_map = {
        str(valor).strip().lower(): idx
        for idx, valor in enumerate(rows[0])
        if valor is not None
    }

    nome_idx = header_map.get(NOME_COL)
    fazenda_idx = header_map.get(FAZENDA_COL)
    peso_idx = header_map.get(PESO_COL)
    tags_idx = header_map.get(TAGS_COL)

    if nome_idx is None or fazenda_idx is None or peso_idx is None:
        raise ValueError(
            "Alguma das colunas esperadas nao foi encontrada na planilha legada."
        )

    linhas: list[IndiceLinha] = []
    for row in rows[1:]:
        linhas.append(
            IndiceLinha(
                nome_arquivo=str(row[nome_idx]),
                fazenda=str(row[fazenda_idx]),
                peso=Decimal(str(row[peso_idx])),
                tags=normalizar_tags(row[tags_idx]) if tags_idx is not None else [],
            )
        )

    return linhas


def inicializar_banco_indice(
    linhas: Iterable[IndiceLinha],
    db_path: str | None = None,
) -> None:
    db_path_resolvido = _resolver_db_path(db_path)
    db_file = Path(db_path_resolvido)
    db_file.parent.mkdir(parents=True, exist_ok=True)
    engine = _engine(str(db_file))
    _criar_schema(engine)

    registros: list[dict[str, Any]] = [
        {
            "ordem": ordem,
            "nome_arquivo": linha.nome_arquivo,
            "fazenda": linha.fazenda,
            "peso": str(linha.peso),
            "tags": ", ".join(linha.tags),
        }
        for ordem, linha in enumerate(linhas)
    ]

    with engine.begin() as conn:
        conn.execute(indice_table.delete())
        if registros:
            conn.execute(
                indice_table.insert(),
                registros,
            )


def garantir_banco_indice(
    db_path: str | None = None,
    legacy_path: str | None = None,
) -> None:
    db_path_resolvido = _resolver_db_path(db_path)
    legacy_path_resolvido = _resolver_legacy_path(legacy_path)
    db_file = Path(db_path_resolvido)
    db_file.parent.mkdir(parents=True, exist_ok=True)
    engine = _engine(str(db_file))
    _criar_schema(engine)

    possui_registros = _contar_registros(engine) > 0
    if possui_registros:
        return

    legacy_file = Path(legacy_path_resolvido)
    if not legacy_file.exists():
        return

    linhas = _carregar_planilha_legada(str(legacy_file))
    inicializar_banco_indice(linhas, db_path=str(db_file))


def carregar_indice(db_path: str = INDICE_DB_PATH) -> list[IndiceLinha]:
    db_path_resolvido = _resolver_db_path(db_path)
    garantir_banco_indice(db_path=db_path_resolvido)
    engine = _engine(db_path_resolvido)

    with engine.connect() as conn:
        rows = conn.execute(
            select(
                indice_table.c.nome_arquivo,
                indice_table.c.fazenda,
                indice_table.c.peso,
                indice_table.c.tags,
            ).order_by(indice_table.c.ordem)
        ).all()

    return [
        IndiceLinha(
            nome_arquivo=row.nome_arquivo,
            fazenda=row.fazenda,
            peso=Decimal(row.peso),
            tags=normalizar_tags(row.tags),
        )
        for row in rows
    ]


def atualizar_tags(nome_arquivo: str, tags: str, db_path: str | None = None) -> None:
    db_path_resolvido = _resolver_db_path(db_path)
    garantir_banco_indice(db_path=db_path_resolvido)
    engine = _engine(db_path_resolvido)

    with engine.begin() as conn:
        result = conn.execute(
            update(indice_table)
            .where(indice_table.c.nome_arquivo == nome_arquivo)
            .values(tags=tags)
        )

    if result.rowcount == 0:
        raise ValueError(f"Imagem '{nome_arquivo}' nao encontrada no indice.")
