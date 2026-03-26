from collections.abc import Iterable

from openpyxl import load_workbook

from src.config import INDICE_PATH, MODELOS_PARA_AVALIACAO, NOME_COL


def construir_colunas_metricas(
    modelos: Iterable[str] | None = None,
) -> list[str]:
    modelos_avaliados = list(modelos) if modelos is not None else list(MODELOS_PARA_AVALIACAO)

    colunas = [
        "area_ground_truth",
        "perimetro_ground_truth",
    ]

    for nome_metrica in ("area", "perimetro", "iou"):
        for modelo in modelos_avaliados:
            colunas.append(f"{nome_metrica}_{modelo}")

    return colunas


def garantir_colunas_metricas_excel(
    indice_path: str = INDICE_PATH,
    modelos: Iterable[str] | None = None,
) -> dict[str, int]:
    """
    Garante que o arquivo Excel de índice possua as colunas de métricas
    esperadas para a avaliação.

    Retorna um mapa ``nome_coluna -> numero_da_coluna`` para facilitar uso
    posterior pela camada que preencherá os valores.
    """
    workbook = load_workbook(indice_path)
    worksheet = workbook.active
    colunas = _garantir_colunas_metricas_worksheet(worksheet, modelos)
    workbook.save(indice_path)
    workbook.close()
    return colunas


def _garantir_colunas_metricas_worksheet(
    worksheet,
    modelos: Iterable[str] | None = None,
) -> dict[str, int]:
    header_map = {
        str(cell.value).strip().lower(): cell.column
        for cell in worksheet[1]
        if cell.value is not None
    }

    for nome_coluna in construir_colunas_metricas(modelos):
        if nome_coluna.lower() in header_map:
            continue

        nova_coluna = worksheet.max_column + 1
        worksheet.cell(row=1, column=nova_coluna, value=nome_coluna)
        header_map[nome_coluna.lower()] = nova_coluna

    return {
        nome_coluna: header_map[nome_coluna.lower()]
        for nome_coluna in construir_colunas_metricas(modelos)
    }


def _resolver_coluna_nome_arquivo(worksheet) -> int:
    header_map = {
        str(cell.value).strip().lower(): cell.column
        for cell in worksheet[1]
        if cell.value is not None
    }
    nome_column = header_map.get(NOME_COL)
    if nome_column is None:
        raise ValueError(
            f"Coluna obrigatória '{NOME_COL}' não encontrada no arquivo Excel."
        )
    return nome_column


def _resolver_linha_por_nome_arquivo(worksheet, nome_arquivo: str, nome_column: int) -> int:
    for row_index in range(2, worksheet.max_row + 1):
        cell_value = worksheet.cell(row=row_index, column=nome_column).value
        if cell_value is None:
            continue
        if str(cell_value).strip() == nome_arquivo:
            return row_index

    raise ValueError(f"Nome de arquivo '{nome_arquivo}' não encontrado no índice Excel.")


def salvar_registro_excel(
    nome_metrica: str,
    nome_arquivo: str,
    resultado: int | float | None,
    modelo: str | None = None,
) -> None:
    raise NotImplementedError("Use a persistência em lote da camada de avaliação.")


def salvar_registros_excel(registros: list[dict[str, str | int | float | None]]) -> None:
    if not registros:
        return

    workbook = load_workbook(INDICE_PATH)
    worksheet = workbook.active

    colunas_metricas = _garantir_colunas_metricas_worksheet(worksheet)
    nome_column = _resolver_coluna_nome_arquivo(worksheet)

    for registro in registros:
        nome_arquivo = registro.get("nome_arquivo")
        if not isinstance(nome_arquivo, str) or not nome_arquivo:
            raise ValueError("Cada registro deve conter 'nome_arquivo' válido.")

        row_index = _resolver_linha_por_nome_arquivo(worksheet, nome_arquivo, nome_column)

        for nome_coluna, valor in registro.items():
            if nome_coluna == "nome_arquivo":
                continue
            column_index = colunas_metricas.get(nome_coluna)
            if column_index is None:
                continue
            worksheet.cell(row=row_index, column=column_index, value=valor)

    workbook.save(INDICE_PATH)
    workbook.close()
