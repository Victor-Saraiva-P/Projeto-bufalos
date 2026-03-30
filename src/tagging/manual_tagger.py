from __future__ import annotations

import os
import sys
import tkinter as tk
from dataclasses import dataclass

if __package__ is None or __package__ == "":
    PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    if PROJECT_ROOT not in sys.path:
        sys.path.insert(0, PROJECT_ROOT)

from openpyxl import load_workbook
from PIL import Image, ImageTk

from src.config import IMAGES_DIR, INDICE_PATH, TAGS_COL
from src.io.indice_loader import carregar_indice_excel
from src.io.path_resolver import PathResolver
from src.models import Imagem


TAG_OPTIONS = [
    "multi_bufalos",
    "cortado",
    "angulo_extremo",
    "baixo_contraste",
    "ocluido",
]
MAX_IMAGE_SIZE = (1100, 700)
BUTTON_DEFAULT_BG = "#d6d3d1"
BUTTON_SELECTED_BG = "#0f766e"
BUTTON_DEFAULT_FG = "#1c1917"
BUTTON_SELECTED_FG = "#f5f5f4"
_PATH_RESOLVER = PathResolver.from_config()


@dataclass(frozen=True)
class PendingImage:
    row_number: int
    line_number: int
    imagem: Imagem


def _resolve_image_path(nome_arquivo: str) -> str:
    caminho_padrao = _PATH_RESOLVER.caminho_foto_original(nome_arquivo)
    if os.path.exists(caminho_padrao):
        return caminho_padrao

    caminho_direto = os.path.join(IMAGES_DIR, nome_arquivo)
    if os.path.exists(caminho_direto):
        return caminho_direto

    raise FileNotFoundError(
        f"Imagem nao encontrada para '{nome_arquivo}'. "
        f"Tentativas: '{caminho_padrao}' e '{caminho_direto}'."
    )


def _is_pending(tags: list[str]) -> bool:
    return len(tags) == 0


class ExcelTagRepository:
    def __init__(self, indice_path: str = INDICE_PATH, tags_col: str = TAGS_COL):
        self.indice_path = indice_path
        self.tags_col = tags_col
        self.workbook = load_workbook(self.indice_path)
        self.worksheet = self.workbook.active
        self.tags_column = self._resolve_tags_column()
        self.pending_images = self._build_pending_images()
        self.current_index = 0

    def _resolve_tags_column(self) -> int:
        header_map = {
            str(cell.value).strip().lower(): cell.column
            for cell in self.worksheet[1]
            if cell.value is not None
        }
        tags_column = header_map.get(self.tags_col)

        if tags_column is None:
            tags_column = self.worksheet.max_column + 1
            self.worksheet.cell(row=1, column=tags_column, value=self.tags_col)
            self.workbook.save(self.indice_path)

        return tags_column

    def _build_pending_images(self) -> list[PendingImage]:
        linhas = carregar_indice_excel()
        return [
            PendingImage(
                row_number=idx,
                line_number=idx - 1,
                imagem=linha,
            )
            for idx, linha in enumerate(linhas, start=2)
            if _is_pending(linha.nomes_tags)
        ]

    def get_current_pending(self) -> PendingImage | None:
        if self.current_index >= len(self.pending_images):
            return None

        return self.pending_images[self.current_index]

    def save_current_and_advance(self, tags_value: str) -> PendingImage | None:
        current_pending = self.get_current_pending()
        if current_pending is None:
            return None

        self.worksheet.cell(
            row=current_pending.row_number,
            column=self.tags_column,
            value=tags_value,
        )
        self.workbook.save(self.indice_path)
        self.current_index += 1
        return self.get_current_pending()

    def close(self) -> None:
        self.workbook.close()


class ManualTaggerApp:
    def __init__(self, master: tk.Tk):
        self.master = master
        self.master.title("Tag de imagens")
        self.master.geometry("1280x920")
        self.master.protocol("WM_DELETE_WINDOW", self.exit_app)

        self.repository = ExcelTagRepository()
        self.selected_tags: set[str] = set()
        self.current_pending: PendingImage | None = None
        self.button_by_tag: dict[str, tk.Button] = {}
        self.tk_image: ImageTk.PhotoImage | None = None
        self.image_available = False

        self._build_layout()
        self._bind_keys()
        self._load_next_pending()

    def _build_layout(self) -> None:
        self.master.columnconfigure(0, weight=1)
        self.master.rowconfigure(1, weight=1)

        top_frame = tk.Frame(self.master, padx=16, pady=12)
        top_frame.grid(row=0, column=0, sticky="ew")
        top_frame.columnconfigure(0, weight=1)

        self.status_label = tk.Label(
            top_frame,
            text="Carregando...",
            anchor="w",
            justify="left",
            font=("Helvetica", 14, "bold"),
        )
        self.status_label.grid(row=0, column=0, sticky="ew")

        self.hint_label = tk.Label(
            top_frame,
            text="Selecione tags com os numeros e confirme com Enter.",
            anchor="w",
            justify="left",
            font=("Helvetica", 11),
        )
        self.hint_label.grid(row=1, column=0, sticky="ew", pady=(4, 0))

        image_frame = tk.Frame(self.master, padx=16, pady=8)
        image_frame.grid(row=1, column=0, sticky="nsew")
        image_frame.columnconfigure(0, weight=1)
        image_frame.rowconfigure(0, weight=1)

        self.image_label = tk.Label(
            image_frame,
            text="",
            bd=1,
            relief="solid",
            anchor="center",
        )
        self.image_label.grid(row=0, column=0, sticky="nsew")

        controls_frame = tk.Frame(self.master, padx=16, pady=16)
        controls_frame.grid(row=2, column=0, sticky="ew")
        controls_frame.columnconfigure(len(TAG_OPTIONS), weight=1)

        for idx, tag in enumerate(TAG_OPTIONS, start=1):
            button = tk.Button(
                controls_frame,
                text=f"{idx}. {tag}",
                width=18,
                padx=10,
                pady=14,
                command=lambda current_tag=tag: self.toggle_tag(current_tag),
            )
            button.grid(row=0, column=idx - 1, padx=6, pady=4, sticky="ew")
            self.button_by_tag[tag] = button

        self.finish_button = tk.Button(
            controls_frame,
            text="Enter. Finalizar",
            width=18,
            padx=10,
            pady=14,
            bg="#2563eb",
            fg="#eff6ff",
            activebackground="#1d4ed8",
            activeforeground="#eff6ff",
            command=self.finish_current_image,
        )
        self.finish_button.grid(
            row=0,
            column=len(TAG_OPTIONS),
            padx=(18, 6),
            pady=4,
            sticky="ew",
        )

        self.exit_button = tk.Button(
            controls_frame,
            text="Sair",
            width=12,
            padx=10,
            pady=14,
            bg="#57534e",
            fg="#fafaf9",
            activebackground="#44403c",
            activeforeground="#fafaf9",
            command=self.exit_app,
        )
        self.exit_button.grid(
            row=0,
            column=len(TAG_OPTIONS) + 1,
            padx=(6, 0),
            pady=4,
            sticky="ew",
        )

        self._refresh_buttons()

    def _bind_keys(self) -> None:
        for idx, tag in enumerate(TAG_OPTIONS, start=1):
            self.master.bind(str(idx), lambda event, current_tag=tag: self.toggle_tag(current_tag))
        self.master.bind("<Return>", lambda event: self.finish_current_image())

    def _refresh_buttons(self) -> None:
        for tag, button in self.button_by_tag.items():
            is_selected = tag in self.selected_tags
            button.configure(
                bg=BUTTON_SELECTED_BG if is_selected else BUTTON_DEFAULT_BG,
                fg=BUTTON_SELECTED_FG if is_selected else BUTTON_DEFAULT_FG,
                activebackground=BUTTON_SELECTED_BG if is_selected else BUTTON_DEFAULT_BG,
                activeforeground=BUTTON_SELECTED_FG if is_selected else BUTTON_DEFAULT_FG,
                relief="sunken" if is_selected else "raised",
            )

    def toggle_tag(self, tag: str) -> None:
        if self.current_pending is None:
            return

        if tag in self.selected_tags:
            self.selected_tags.remove(tag)
        else:
            self.selected_tags.add(tag)

        self._refresh_buttons()

    def finish_current_image(self) -> None:
        if self.current_pending is None or not self.image_available:
            return

        self._save_current_image()
        self._load_next_pending()

    def exit_app(self) -> None:
        if self.current_pending is not None and self.image_available and self.selected_tags:
            self._save_current_image()

        self.repository.close()
        self.master.destroy()

    def _build_tags_value(self) -> str:
        selected = [tag for tag in TAG_OPTIONS if tag in self.selected_tags]
        if not selected:
            return "ok"

        return ", ".join(selected)

    def _save_current_image(self) -> None:
        if self.current_pending is None or not self.image_available:
            return

        tags_value = self._build_tags_value()
        self.current_pending = self.repository.save_current_and_advance(tags_value)

    def _load_next_pending(self) -> None:
        self.selected_tags.clear()
        self.image_available = False
        self._refresh_buttons()

        if self.current_pending is None:
            self.current_pending = self.repository.get_current_pending()

        pending = self.current_pending

        if pending is None:
            self.status_label.configure(text="Nao ha mais imagens para tag-acao.")
            self.hint_label.configure(text="Todas as linhas da coluna tags ja foram preenchidas.")
            self.image_label.configure(image="", text="Fila concluida.")
            self.finish_button.configure(state="disabled")
            self.exit_button.configure(state="normal")
            return

        self.finish_button.configure(state="normal")
        linha = pending.imagem
        self.status_label.configure(
            text=(
                f"Linha {pending.line_number}: {linha.nome_arquivo} | "
                f"fazenda={linha.fazenda} | peso={linha.peso}"
            )
        )
        self.hint_label.configure(
            text="Enter sem selecao grava ok. Numeros alternam as tags."
        )

        try:
            image_path = _resolve_image_path(linha.nome_arquivo)
            self._display_image(image_path)
            self.finish_button.configure(state="normal")
        except FileNotFoundError as exc:
            self.tk_image = None
            self.finish_button.configure(state="disabled")
            self.image_label.configure(
                image="",
                text=str(exc),
                justify="center",
                wraplength=1000,
            )

    def _display_image(self, image_path: str) -> None:
        image = Image.open(image_path)
        image.thumbnail(MAX_IMAGE_SIZE)
        self.tk_image = ImageTk.PhotoImage(image)
        self.image_available = True
        self.image_label.configure(image=self.tk_image, text="")


def main() -> None:
    root = tk.Tk()
    ManualTaggerApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
